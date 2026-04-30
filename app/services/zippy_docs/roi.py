"""ROI Analysis generator — fills Beacon's ROI Excel template.

The template (``Beacon_ROI_Template.xlsx``) is the source of truth for
ROI math. Three sheets are pure formulas (Executive Summary, Man-Hour
Model, ROI Analysis) and must never be touched. Zippy fills only:

  * ``Survey Input`` — raw Q&A answers pasted verbatim into column C.
  * ``1. Inputs & Assumptions`` — parsed numeric model values in C4..C20.

Every other cell (and any cell whose value starts with ``=``) is left
exactly as the template had it, so the formulas drive everything when
Google Sheets opens the file.
"""
from __future__ import annotations

import io
import json
import logging
from dataclasses import dataclass
from datetime import datetime
from typing import Optional

from sqlmodel import select as sm_select

from app.clients.anthropic_client import get_anthropic_client
from app.clients.google_drive import upload_as_google_sheet
from app.config import settings
from app.database import AsyncSessionLocal as async_session
from app.models.user_email_connection import UserEmailConnection
from app.models.zippy import IndexedDriveFile
from app.services.zippy_docs.base import (
    GeneratedDocument,
    build_output_path,
    human_today,
)

logger = logging.getLogger(__name__)


# ── Input dataclass ─────────────────────────────────────────────────────────


@dataclass
class ROIInput:
    """Survey responses + optional pre-parsed numeric values for the ROI model.

    Raw ``q*`` strings come straight off the AE's form-response email —
    we hand them to Claude to parse into the numeric fields below. If the
    caller already has parsed numbers, they win and we skip the LLM hop.
    """
    client_name: str
    prepared_by: str = "Beacon"
    report_date: Optional[str] = None  # e.g. "April 2026"

    # Raw Google Form responses
    q1_reason: Optional[str] = None
    q2_impls_per_year: Optional[str] = None
    q3_team_size: Optional[str] = None
    q4_ftes_per_impl: Optional[str] = None
    q5_duration_range: Optional[str] = None
    q6_inception_weeks: Optional[str] = None
    q7_solutioning_weeks: Optional[str] = None
    q8_config_weeks: Optional[str] = None
    q9_data_migration_weeks: Optional[str] = None
    q10_testing_weeks: Optional[str] = None
    q11_cutover_weeks: Optional[str] = None
    q12_fte_cost_usd: Optional[str] = None
    q13_ramp_up: Optional[str] = None
    q14_new_headcount: Optional[str] = None

    # Parsed numeric values — used directly if provided, else Claude parses.
    impls_per_year: Optional[int] = None
    team_ftes: Optional[int] = None
    ftes_per_impl: Optional[float] = None
    fte_cost_usd: Optional[float] = None
    new_headcount: Optional[int] = None
    inception_weeks: Optional[float] = None
    solutioning_weeks: Optional[float] = None
    config_weeks: Optional[float] = None
    data_migration_weeks: Optional[float] = None
    testing_weeks: Optional[float] = None
    cutover_weeks: Optional[float] = None


# ── Drive template discovery ────────────────────────────────────────────────
#
# Same shape as mom.py and nda.py: scan IndexedDriveFile for keyword matches,
# prefer the user's own copy then admin/shared, never leak someone else's
# private file. Order matters — most specific keyword first.

_ROI_KEYWORDS = [
    "beacon_roi_template",
    "beacon roi template",
    "roi template",
    "roi model",
    "roi analysis",
    "roi",
]


async def _find_roi_template_row(
    user_id: Optional[str] = None,
) -> Optional[IndexedDriveFile]:
    """Locate the indexed ROI template row, preferring user's own then admin."""
    from sqlalchemy import or_

    async with async_session() as session:
        for keyword in _ROI_KEYWORDS:
            stmt = sm_select(IndexedDriveFile).where(
                IndexedDriveFile.name.ilike(f"%{keyword}%"),
            )
            if user_id is not None:
                stmt = stmt.where(
                    or_(
                        IndexedDriveFile.owner_user_id == user_id,
                        IndexedDriveFile.is_admin == True,  # noqa: E712
                    )
                )
            else:
                stmt = stmt.where(IndexedDriveFile.is_admin == True)  # noqa: E712

            if user_id is not None:
                from sqlalchemy import case
                user_priority = case(
                    (IndexedDriveFile.owner_user_id == user_id, 0),
                    else_=1,
                )
                stmt = stmt.order_by(
                    user_priority,
                    IndexedDriveFile.last_indexed_at.desc(),
                )
            else:
                stmt = stmt.order_by(IndexedDriveFile.last_indexed_at.desc())

            row = (await session.execute(stmt.limit(1))).scalar_one_or_none()
            if row:
                logger.info(
                    "ROI template found via keyword=%r: name=%s file_id=%s "
                    "mime=%s owner=%s is_admin=%s",
                    keyword, row.name, row.drive_file_id, row.mime_type,
                    row.owner_user_id, row.is_admin,
                )
                return row
    logger.info("No ROI template matched for user_id=%s", user_id)
    return None


async def _fetch_template_bytes(
    user_id: Optional[str] = None,
) -> tuple[Optional[bytes], Optional[str]]:
    """Download the ROI template bytes for a specific user.

    Native Google Sheets files are exported as .xlsx so openpyxl can read
    them. Real .xlsx files are downloaded directly.
    """
    row = await _find_roi_template_row(user_id=user_id)
    if not row:
        return None, (
            "No ROI template found in Drive. Make sure "
            "'Beacon_ROI_Template.xlsx' is in your indexed Drive folder "
            "and re-sync."
        )

    async with async_session() as session:
        result = await session.execute(
            sm_select(UserEmailConnection).where(
                UserEmailConnection.user_id == row.owner_user_id,
                UserEmailConnection.is_active == True,  # noqa: E712
            )
        )
        connection = result.scalar_one_or_none()

    if not connection:
        return None, (
            f"No active Drive connection for the owner of '{row.name}'. "
            "Reconnect Google in Settings and re-sync."
        )

    try:
        if row.mime_type == "application/vnd.google-apps.spreadsheet":
            import httpx
            from app.clients.google_drive import _ensure_token, DRIVE_API_BASE
            access_token, _ = await _ensure_token(
                connection.token_data,
                settings.gmail_client_id,
                settings.gmail_client_secret,
            )
            xlsx_mime = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            async with httpx.AsyncClient(timeout=60) as http:
                resp = await http.get(
                    f"{DRIVE_API_BASE}/files/{row.drive_file_id}/export",
                    headers={"Authorization": f"Bearer {access_token}"},
                    params={"mimeType": xlsx_mime},
                )
            if resp.status_code == 200:
                return resp.content, None
            return None, (
                f"Drive export failed for '{row.name}' (HTTP "
                f"{resp.status_code}): {resp.text[:300]}"
            )

        from app.clients import google_drive
        data, _mime, _updated = await google_drive.download_file_bytes(
            file_id=row.drive_file_id,
            mime_type=row.mime_type,
            token_data=connection.token_data,
            client_id=settings.gmail_client_id,
            client_secret=settings.gmail_client_secret,
        )
        if not data:
            return None, f"Drive returned empty bytes for '{row.name}'."
        return data, None
    except Exception as exc:
        logger.exception("Failed to download ROI template '%s': %s", row.name, exc)
        return None, f"Drive download failed for '{row.name}': {exc}"


# ── Claude value parser ─────────────────────────────────────────────────────


async def _parse_values_with_claude(
    data: ROIInput,
    client,
    model: str,
) -> dict:
    """Parse raw Q&A responses into numeric model values.

    Returns a dict with keys matching the parsed-value fields on ROIInput.
    Empty dict if Claude is unavailable, the prompt has nothing to parse,
    or parsing fails — the caller falls back to whatever numeric fields
    the AE provided directly.
    """
    if not client:
        return {}

    raw_responses = {
        "Q2 (implementations/year)": data.q2_impls_per_year,
        "Q3 (team size FTEs)": data.q3_team_size,
        "Q4 (FTEs per impl)": data.q4_ftes_per_impl,
        "Q6 (inception weeks)": data.q6_inception_weeks,
        "Q7 (solutioning weeks)": data.q7_solutioning_weeks,
        "Q8 (config weeks)": data.q8_config_weeks,
        "Q9 (data migration weeks)": data.q9_data_migration_weeks,
        "Q10 (testing weeks)": data.q10_testing_weeks,
        "Q11 (cutover weeks)": data.q11_cutover_weeks,
        "Q12 (FTE cost USD/year)": data.q12_fte_cost_usd,
        "Q14 (new headcount)": data.q14_new_headcount,
    }
    raw_responses = {k: v for k, v in raw_responses.items() if v}
    if not raw_responses:
        return {}

    prompt = f"""Extract numeric values from these Google Form survey responses.

Survey responses:
{json.dumps(raw_responses, indent=2)}

For each field, return the numeric value only (no units, no currency symbols).
Rules:
- Q2: return the FULL-MODULE count only (e.g. "700 total, 400 full module" -> 400)
- Q3: total team FTEs (e.g. "24" -> 24)
- Q4: FTEs per single implementation (e.g. "3" -> 3.0)
- Q6-Q11: use the MIDPOINT of any range given (e.g. "1-4 weeks" -> 2.0, "4 weeks" -> 4.0)
         return 0 if the answer is "ongoing" or "not discrete"
- Q12: annual cost in USD only as a number (e.g. "$40,000" -> 40000, "£65,000" -> convert to USD at 1.25)
- Q14: net new FTEs as integer (e.g. "Net 0" -> 0, "+3" -> 3, "No change" -> 0)

Return ONLY valid JSON with these exact keys (omit keys where value is unknown):
{{
  "impls_per_year": <int>,
  "team_ftes": <int>,
  "ftes_per_impl": <float>,
  "fte_cost_usd": <float>,
  "new_headcount": <int>,
  "inception_weeks": <float>,
  "solutioning_weeks": <float>,
  "config_weeks": <float>,
  "data_migration_weeks": <float>,
  "testing_weeks": <float>,
  "cutover_weeks": <float>
}}"""

    try:
        response = await client.messages.create(
            model=model,
            max_tokens=500,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = "".join(
            b.text for b in response.content
            if getattr(b, "type", "") == "text"
        )
        start, end = raw.find("{"), raw.rfind("}")
        if start == -1 or end == -1:
            return {}
        parsed = json.loads(raw[start: end + 1])
        return parsed if isinstance(parsed, dict) else {}
    except Exception as exc:
        logger.warning("ROI value parsing failed: %s", exc)
        return {}


# ── Excel cell filler ───────────────────────────────────────────────────────


def _is_formula(cell) -> bool:
    """True if a cell holds a formula — never overwrite these."""
    return isinstance(cell.value, str) and cell.value.startswith("=")


def _fill_template(xlsx_bytes: bytes, data: ROIInput, parsed: dict) -> bytes:
    """Fill exactly the known input cells in the ROI template.

    Touches ``Survey Input`` (raw Q&A) and ``1. Inputs & Assumptions``
    (numeric model values). Every formula cell is skipped — a defensive
    guard against the template being rearranged. All ROI math runs from
    formulas in the other three sheets, which we leave untouched.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=False)

    # ── Sheet: Survey Input ────────────────────────────────────────────────
    if "Survey Input" in wb.sheetnames:
        si = wb["Survey Input"]

        def fill_cell_if_placeholder(ws, placeholder: str, new_value: str) -> None:
            """Replace the first cell whose text contains ``placeholder``.

            Walks the whole sheet — the template positions these client
            details freely (header rows, banner cells, etc.), so we don't
            hardcode coordinates. Skips formula cells defensively.
            """
            for row in ws.iter_rows():
                for cell in row:
                    if (cell.value
                            and isinstance(cell.value, str)
                            and placeholder in cell.value
                            and not _is_formula(cell)):
                        cell.value = new_value
                        return

        fill_cell_if_placeholder(si, "[Enter Client Name]", data.client_name)
        fill_cell_if_placeholder(
            si, "[Month Year]", data.report_date or human_today()
        )
        fill_cell_if_placeholder(si, "[Your Name]", data.prepared_by)

        # Q1..Q14 raw answers — paste verbatim into the "Client Answer"
        # column (column C). The template has the question key (e.g. "Q1")
        # in column A, so we walk rows looking for a Q-key in col A and
        # write into col C of the same row.
        q_map = {
            "Q1": data.q1_reason,
            "Q2": data.q2_impls_per_year,
            "Q3": data.q3_team_size,
            "Q4": data.q4_ftes_per_impl,
            "Q5": data.q5_duration_range,
            "Q6": data.q6_inception_weeks,
            "Q7": data.q7_solutioning_weeks,
            "Q8": data.q8_config_weeks,
            "Q9": data.q9_data_migration_weeks,
            "Q10": data.q10_testing_weeks,
            "Q11": data.q11_cutover_weeks,
            "Q12": data.q12_fte_cost_usd,
            "Q13": data.q13_ramp_up,
            "Q14": data.q14_new_headcount,
        }
        for row in si.iter_rows():
            first_cell = row[0].value if row else None
            if first_cell and isinstance(first_cell, str):
                q_key = first_cell.strip()
                if q_key in q_map and q_map[q_key]:
                    answer_cell = row[2] if len(row) > 2 else None
                    if answer_cell and not _is_formula(answer_cell):
                        answer_cell.value = q_map[q_key]

    # ── Sheet: 1. Inputs & Assumptions ─────────────────────────────────────
    inputs_sheet_name: Optional[str] = None
    for name in wb.sheetnames:
        lname = name.lower()
        if "inputs" in lname or lname.startswith("1."):
            inputs_sheet_name = name
            break

    if inputs_sheet_name:
        inp = wb[inputs_sheet_name]

        def safe_set(ws, row_num: int, col_num: int, value) -> None:
            """Set cell only if it's not a formula and value is not None."""
            if value is None:
                return
            cell = ws.cell(row=row_num, column=col_num)
            if _is_formula(cell):
                return
            cell.value = value

        # Use parsed values from Claude first; fall back to direct fields
        # on the dataclass so callers that already have numbers can skip
        # the LLM hop entirely.
        p = parsed

        safe_set(inp, 4,  3, p.get("impls_per_year")       or data.impls_per_year)
        safe_set(inp, 5,  3, p.get("team_ftes")            or data.team_ftes)
        safe_set(inp, 6,  3, p.get("ftes_per_impl")        or data.ftes_per_impl)
        safe_set(inp, 7,  3, p.get("fte_cost_usd")         or data.fte_cost_usd)
        safe_set(inp, 8,  3, p.get("new_headcount")        or data.new_headcount)
        # C9 (working hrs/week) and C10 (weeks/year) intentionally untouched —
        # template defaults (40h, 52w) are the right answer 99% of the time.
        safe_set(inp, 15, 3, p.get("inception_weeks")      or data.inception_weeks)
        safe_set(inp, 16, 3, p.get("solutioning_weeks")    or data.solutioning_weeks)
        safe_set(inp, 17, 3, p.get("config_weeks")         or data.config_weeks)
        safe_set(inp, 18, 3, p.get("data_migration_weeks") or data.data_migration_weeks)
        safe_set(inp, 19, 3, p.get("testing_weeks")        or data.testing_weeks)
        safe_set(inp, 20, 3, p.get("cutover_weeks")        or data.cutover_weeks)

        for row in inp.iter_rows(max_row=3):
            for cell in row:
                if (cell.value
                        and isinstance(cell.value, str)
                        and "[Client Name]" in cell.value
                        and not _is_formula(cell)):
                    cell.value = cell.value.replace(
                        "[Client Name]", data.client_name
                    )

    # Executive Summary banner — replace [Client Name]/[Month Year] placeholders
    # in the title rows only. We never touch the formula-driven rows below.
    if "Executive Summary" in wb.sheetnames:
        es = wb["Executive Summary"]
        for row in es.iter_rows(max_row=3):
            for cell in row:
                if (cell.value
                        and isinstance(cell.value, str)
                        and not _is_formula(cell)):
                    if "[Client Name]" in cell.value:
                        cell.value = cell.value.replace(
                            "[Client Name]", data.client_name
                        )
                    if "[Month Year]" in cell.value:
                        cell.value = cell.value.replace(
                            "[Month Year]",
                            data.report_date or human_today(),
                        )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Fallback xlsx (template not in Drive) ───────────────────────────────────


def _render_fallback_xlsx(data: ROIInput, path) -> None:
    """Minimal one-sheet summary so the tool never hard-fails."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "ROI Summary"
    ws["A1"] = f"ROI Analysis — {data.client_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        "Note: Beacon_ROI_Template.xlsx was not found in Drive. "
        "Add it to your indexed folder."
    )
    ws["A2"].font = Font(italic=True, color="FF0000")

    row = 4
    fields = [
        ("Client", data.client_name),
        ("Implementations/Year", data.impls_per_year),
        ("Team FTEs", data.team_ftes),
        ("FTEs per Impl", data.ftes_per_impl),
        ("FTE Cost USD/yr", data.fte_cost_usd),
        ("New Headcount", data.new_headcount),
        ("Inception Weeks", data.inception_weeks),
        ("Solutioning Weeks", data.solutioning_weeks),
        ("Config Weeks", data.config_weeks),
        ("Data Migration Weeks", data.data_migration_weeks),
        ("Testing Weeks", data.testing_weeks),
        ("Cutover Weeks", data.cutover_weeks),
    ]
    for label, value in fields:
        if value is not None:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1

    wb.save(str(path))


# ── Drive upload ────────────────────────────────────────────────────────────


async def _try_upload_to_sheets(
    doc: GeneratedDocument,
    path,
    client_name: str,
    *,
    user_id: Optional[str],
) -> None:
    """Upload the filled .xlsx to Drive as an editable Google Sheet."""
    if doc.drive_url:
        return
    if user_id:
        from app.services.zippy_docs.base import (
            cache_upload,
            get_cached_upload,
        )
        cached = get_cached_upload(str(user_id), client_name, doc.kind)
        if cached:
            doc.drive_url = cached
            logger.info("Reusing cached ROI Sheets upload: %s", cached)
            return
    try:
        from sqlalchemy import or_, case as sa_case

        async with async_session() as session:
            stmt = sm_select(UserEmailConnection).where(
                UserEmailConnection.is_active == True,  # noqa: E712
            )
            if user_id:
                stmt = stmt.where(
                    or_(
                        UserEmailConnection.user_id == user_id,
                        UserEmailConnection.is_admin_folder == True,  # noqa: E712
                    )
                )
                priority = sa_case(
                    (UserEmailConnection.user_id == user_id, 0), else_=1
                )
                stmt = stmt.order_by(priority)
            result = await session.execute(stmt.limit(1))
            connection = result.scalar_one_or_none()

        if not connection:
            logger.info("No active Drive connection — skipping ROI sheet upload")
            return

        xlsx_bytes = path.read_bytes()
        sheet_name = (
            f"ROI — {client_name} — {doc.created_at.strftime('%d %b %Y')}"
        )
        folder_id = getattr(connection, "selected_drive_folder_id", None)

        file_id, web_view_link = await upload_as_google_sheet(
            filename=sheet_name,
            docx_bytes=xlsx_bytes,
            token_data=connection.token_data,
            client_id=settings.gmail_client_id,
            client_secret=settings.gmail_client_secret,
            parent_folder_id=folder_id,
        )
        doc.drive_file_id = file_id
        doc.drive_url = web_view_link
        logger.info("ROI uploaded to Google Sheets: %s", web_view_link)
        if user_id and web_view_link:
            from app.services.zippy_docs.base import cache_upload
            cache_upload(str(user_id), client_name, doc.kind, web_view_link)
    except PermissionError as exc:
        logger.info("drive.file scope not yet granted — skipping ROI upload: %s", exc)
    except Exception as exc:
        logger.warning("Google Sheets upload failed for ROI (non-fatal): %s", exc)


# ── Public entry points ─────────────────────────────────────────────────────


async def inspect_roi_template(user_id: Optional[str] = None) -> dict:
    """Confirm the ROI template is reachable; report the survey fields needed."""
    template_bytes, err = await _fetch_template_bytes(user_id=user_id)
    if not template_bytes:
        return {
            "found": False,
            "error": err or "Unknown error fetching ROI template.",
        }

    row = await _find_roi_template_row(user_id=user_id)
    return {
        "found": True,
        "template_name": row.name if row else "Beacon_ROI_Template.xlsx",
        "sheets": [
            "Executive Summary",
            "Survey Input",
            "1. Inputs & Assumptions",
            "2. Man-Hour Model",
            "3. ROI Analysis",
        ],
        "input_fields": [
            "Q2: implementations per year (full-module only)",
            "Q3: implementation team size (FTEs)",
            "Q4: FTEs per single implementation",
            "Q6: Inception/Discovery weeks",
            "Q7: Solutioning/BRD weeks",
            "Q8: Configuration weeks",
            "Q9: Data Migration weeks",
            "Q10: Testing/UAT weeks",
            "Q11: Cutover weeks",
            "Q12: fully-loaded annual FTE cost (USD)",
            "Q14: new headcount planned",
        ],
        "note": (
            "Fill Survey Input sheet with raw answers and "
            "1. Inputs & Assumptions with parsed values. "
            "All ROI calculations are formula-driven — no manual "
            "calculation needed."
        ),
    }


async def generate(
    data: ROIInput,
    user_id: Optional[str] = None,
) -> GeneratedDocument:
    """Generate an ROI Analysis xlsx and upload it as a live Google Sheet."""
    path, url = build_output_path("ROI", data.client_name, extension="xlsx")
    template_bytes, err = await _fetch_template_bytes(user_id=user_id)
    source = "template"

    if not template_bytes:
        logger.warning("ROI template unavailable (%s) — rendering fallback xlsx", err)
        _render_fallback_xlsx(data, path)
        source = "fallback"
    else:
        client = get_anthropic_client()
        parsed = await _parse_values_with_claude(
            data, client, settings.CLAUDE_MODEL_STANDARD
        )
        out_bytes = _fill_template(template_bytes, data, parsed)
        path.write_bytes(out_bytes)

    fields_filled = sum(
        1 for v in [
            data.impls_per_year, data.team_ftes, data.ftes_per_impl,
            data.fte_cost_usd, data.inception_weeks, data.solutioning_weeks,
            data.config_weeks, data.data_migration_weeks,
            data.testing_weeks, data.cutover_weeks,
        ] if v is not None
    )

    result = GeneratedDocument(
        filename=path.name,
        path=str(path),
        url=url,
        kind="roi",
        summary=(
            f"ROI Analysis for {data.client_name} — {source}. "
            f"{fields_filled} input fields populated. "
            "Open the Google Sheet to see live calculated ROI — all "
            "formulas auto-update."
        ),
        created_at=datetime.utcnow(),
    )

    await _try_upload_to_sheets(
        result, path, data.client_name, user_id=user_id
    )
    return result
